"""Aggregate per-sub-capability license flags for unified connectors.

For every connector in the migration pipeline, this script enumerates each
integration's capabilities (via :func:`capabilities_collector.collect_capabilities`)
and emits a row per ``(connector, capability, sub_capability)`` triple. The
license flags for each sub-capability are resolved from the integration's row
in the license-filtering CSV, with a ``supportedModules`` fallback.

Output is written in two formats:
  - **CSV** with grouped continuation rows (the ``Connector ID`` cell is blank
    after the first row of a connector, and the ``Capability`` cell is blank
    after the first row of a capability).
  - **XLSX** with the same data but real *merged cells* for the ``Connector
    ID`` and ``Capability`` columns.

Header layout mirrors ``output_example - Pack and integrations filtering.csv``:

    Connector ID,Capability,sub_capability,cloud_posture,cloud,
    cloud_runtime_security,edr,asm,cloud_appsec,xsiam,exposure_management,
    agentix,tim

Conventions (shared with manifest_generator):
  - capability id: canonical id from ``slugify_capability_name`` (e.g.
    ``Automation`` -> ``automation-and-remediation``).
  - integration id slug: ``commonfields.id`` lowercased with internal
    whitespace runs collapsed to single dashes.
  - sub_capability id: ``<capability_id>_<integration-id-slug>``.

License resolution (per spec), in order:
  1. License CSV row matched by ``commonfields.id``.
  2. License CSV row matched by the integration's display ``name``
     (== pipeline ``Integration ID``).
  3. ``supportedModules`` from the integration YML — each license column is
     TRUE when its column name appears in the module list, else FALSE.
  4. ``supported_modules`` from the parent pack's ``pack_metadata.json`` —
     same 1:1 column-name mapping.
  5. Otherwise the integration is reported as "missing" and excluded.

Usage:
    python -m connectus.connectus_migration.license_aggregator \\
        --pipeline-csv connectus/connectus-migration-pipeline.csv \\
        --license-csv "connectus/Content Packs filtering per license - July 2026 - Pack and integrations filtering.csv" \\
        --output connectus/connector_capability_licenses.csv
"""

import csv
import json
import logging
import re
from pathlib import Path

import typer
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from capabilities_collector import collect_capabilities

try:
    # slugify_capability_name maps a capability bucket key to its canonical id.
    from manifest_generator import slugify_capability_name
except Exception:  # pragma: no cover - fallback when import path differs
    from connectus.connectus_migration.manifest_generator import (  # type: ignore
        slugify_capability_name,
    )

logger = logging.getLogger(__name__)

main = typer.Typer()

# Sentinel used in the license CSV's ``integration name`` column for the
# pack-level aggregate row, which must be ignored for per-integration lookups.
ALL_PACK_SENTINEL = "All pack"

# Fixed columns that precede the license columns in the OUTPUT file.
OUTPUT_LEADING_COLUMNS = ["Connector ID", "Capability", "sub_capability"]

# Number of leading columns in the LICENSE csv ("pack name", "integration name").
LICENSE_LEADING_COLUMNS = 2

# Boolean string literals used throughout the license sheets.
TRUE_STR = "TRUE"
FALSE_STR = "FALSE"
ALL_MODULES = ["cloud_posture", "cloud", "cloud_runtime_security", "edr", "asm", "xsiam", "cloud_appsec", "tim", "agentix", "exposure_management"]

def integration_id_to_slug(integration_id: str) -> str:
    """Lowercase + dash-slug an integration id (``commonfields.id``).

    Internal whitespace runs collapse to a single dash. This mirrors the slug
    embedded in the handler id by ``manifest_generator.derive_handler_id`` so
    the sub-capability ids produced here line up with the manifest output.

    Examples:
        integration_id_to_slug("Salesforce") -> "salesforce"
        integration_id_to_slug("Hello World IAM") -> "hello-world-iam"
        integration_id_to_slug("EWS v2") -> "ews-v2"
    """
    return re.sub(r"\s+", "-", integration_id.strip().lower())


def _select_named_license_columns(
    header: list[str],
    license_leading_columns=LICENSE_LEADING_COLUMNS
) -> tuple[list[str], list[int]]:
    """From the full license header, return the named license columns and the
    indices (into a full data row) they map to.

    Drops the two leading id columns AND any blank-named license columns (the
    output example no longer carries the unnamed columns).
    """
    names: list[str] = []
    indices: list[int] = []
    for idx in range(license_leading_columns, len(header)):
        col_name = header[idx].strip()
        if not col_name:
            continue  # skip unnamed/blank license columns
        names.append(col_name)
        indices.append(idx)
    return names, indices


def load_license_table(
    license_csv: Path,
) -> tuple[list[str], dict[str, list[str]]]:
    """Load the license CSV.

    Returns ``(license_columns, by_name)`` where:
      - ``license_columns`` is the list of NAMED license-column headers
        (blank/unnamed columns dropped).
      - ``by_name`` maps each integration name (excluding ``All pack`` rows)
        to its list of license-flag values sliced to those named columns.

    Last-seen row wins on duplicate integration names (rare; logged).
    """
    with open(license_csv, newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"License CSV {license_csv} is empty.")

        license_columns, col_indices = _select_named_license_columns(header)
        by_name: dict[str, list[str]] = {}
        for row in reader:
            if not row or len(row) < LICENSE_LEADING_COLUMNS:
                continue
            integration_name = row[1].strip()
            if not integration_name or integration_name == ALL_PACK_SENTINEL:
                continue
            values = [
                row[idx].strip() if idx < len(row) else "" for idx in col_indices
            ]
            if integration_name in by_name:
                logger.warning(
                    f"[license_aggregator] Duplicate integration name "
                    f"'{integration_name}' in license CSV; last row wins."
                )
            by_name[integration_name] = values

    return license_columns, by_name


def prase_licenses_to_json(
    license_csv: Path,
):
    """Load the license CSV.

    Returns ``(license_columns, by_name)`` where:
      - ``license_columns`` is the list of NAMED license-column headers
        (blank/unnamed columns dropped).
      - ``by_name`` maps each integration name (excluding ``All pack`` rows)
        to its list of license-flag values sliced to those named columns.

    Last-seen row wins on duplicate integration names (rare; logged).
    """
    licenses = {}
    with open(license_csv, newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"License CSV {license_csv} is empty.")

        for row in reader:
            if not row or len(row) < 3:
                continue
            sub_capability_id = row[2].strip()
            sub_cap_licenses = []
            for i, val in enumerate(row):
                if val == "TRUE":
                    sub_cap_licenses.append(header[i])
            licenses[sub_capability_id] = sub_cap_licenses
    path = Path("connectus/connectus_migration/sub_capabilities_to_licenses.json")
    path.write_text(json.dumps(licenses, indent=2, sort_keys=True))


def load_integration_yml(path: Path) -> dict:
    """Load an integration YML file into a dict (empty dict on missing keys)."""
    with open(path) as fh:
        return yaml.safe_load(fh) or {}


def _modules_to_license_row(
    modules: list[str],
    license_columns: list[str],
    unmatched: set[str],
) -> list[str]:
    """Map a ``supportedModules`` list onto the license columns.

    A column is ``TRUE`` when its (case-insensitive) name appears in
    ``modules``, otherwise ``FALSE``. Any module value that does not match a
    license column is added to ``unmatched`` for end-of-run reporting.
    """
    module_set = {m.strip().lower() for m in modules if m and m.strip()}
    column_set = {c.strip().lower() for c in license_columns}
    for module in module_set:
        if module not in column_set:
            unmatched.add(module)
    return [
        TRUE_STR if col.strip().lower() in module_set else FALSE_STR
        for col in license_columns
    ]


def _supported_modules_from_pack(integration_path: Path) -> list[str]:
    """Read ``supported_modules`` from the integration's pack_metadata.json.

    Returns an empty list when the file or field is absent/invalid.
    """
    pack_root = integration_path.parent.parent.parent
    metadata_path = pack_root / "pack_metadata.json"
    if not metadata_path.is_file():
        return []
    try:
        with open(metadata_path) as fh:
            data = json.load(fh)
    except Exception as exc:
        logger.warning(
            f"[license_aggregator] Failed to parse {metadata_path}: {exc}."
        )
        return []
    modules = data.get("supported_modules", []) or data.get("supportedModules", [])
    return modules if isinstance(modules, list) else []


def resolve_license_row(
    integration_yml: dict,
    integration_path: Path,
    integration_id: str,
    integration_name: str,
    by_name: dict[str, list[str]],
    license_columns: list[str],
    unmatched_modules: set[str],
) -> list[str] | None:
    """Resolve an integration's license row using the full fallback chain.

    Order: license CSV by id -> by name -> integration ``supportedModules``
    -> pack ``supported_modules``. Returns ``None`` when nothing resolves.
    """
    # 1 + 2: direct license-CSV hit (id first, then display name).
    if integration_id and integration_id in by_name:
        return by_name[integration_id]
    if integration_name and integration_name in by_name:
        return by_name[integration_name]

    # 3: integration YML supportedModules.
    yml_modules = integration_yml.get("supportedModules")
    if isinstance(yml_modules, list) and yml_modules:
        return _modules_to_license_row(
            yml_modules, license_columns, unmatched_modules
        )

    # 4: pack_metadata.json supported_modules.
    pack_modules = _supported_modules_from_pack(integration_path)
    if pack_modules:
        return _modules_to_license_row(
            pack_modules, license_columns, unmatched_modules
        )
        
    # 5: unresolved - takes all
    return _modules_to_license_row(
            ALL_MODULES, license_columns, unmatched_modules
        )


def read_pipeline_rows(pipeline_csv: Path) -> list[dict[str, str]]:
    """Read the pipeline CSV into a list of dict rows (DictReader)."""
    with open(pipeline_csv, newline="") as fh:
        reader = csv.DictReader(fh)
        return [row for row in reader]


def group_by_connector(
    rows: list[dict[str, str]],
) -> dict[str, list[dict[str, str]]]:
    """Group pipeline rows by ``Connector ID`` preserving first-appearance order."""
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        connector_id = (row.get("Connector ID") or "").strip()
        if not connector_id:
            continue
        grouped.setdefault(connector_id, []).append(row)
    return grouped


def build_connector_records(
    connector_rows: list[dict[str, str]],
    by_name: dict[str, list[str]],
    license_columns: list[str],
    repo_root: Path,
    missing: list[str],
    unmatched_modules: set[str],
) -> list[tuple[str, str, list[str]]]:
    """Build the (capability_id, sub_capability_id, license_values) records for
    a single connector.

    Capabilities are ordered by first appearance across the connector's
    integrations; within a capability, sub-capabilities follow the order their
    integrations appear in the pipeline. ``missing`` is appended in place with
    the identifier of any integration that has no resolvable license row.
    """
    # capability_id -> ordered list of (sub_cap_id, license_values)
    cap_to_subcaps: dict[str, list[tuple[str, list[str]]]] = {}

    for row in connector_rows:
        pipeline_int_id = (row.get("Integration ID") or "").strip()
        int_path = (row.get("Integration File Path") or "").strip()
        if not int_path:
            logger.warning(
                f"[license_aggregator] Missing Integration File Path for "
                f"'{pipeline_int_id}'; skipping."
            )
            continue

        yml_path = repo_root / int_path
        if not yml_path.is_file():
            logger.warning(
                f"[license_aggregator] Integration yml not found at {yml_path}; "
                f"skipping '{pipeline_int_id}'."
            )
            missing.append(pipeline_int_id or int_path)
            continue

        integration_yml = load_integration_yml(yml_path)
        commonfields_id = (
            integration_yml.get("commonfields", {}).get("id") or ""
        ).strip()
        # Display name; fall back to the pipeline's Integration ID column.
        integration_name = (
            integration_yml.get("name") or pipeline_int_id
        ).strip()

        license_values = resolve_license_row(
            integration_yml,
            yml_path,
            commonfields_id,
            integration_name,
            by_name,
            license_columns,
            unmatched_modules,
        )
        if license_values is None:
            missing.append(commonfields_id or integration_name or pipeline_int_id)
            continue

        capabilities = collect_capabilities(integration_yml)
        if not capabilities:
            continue

        slug = integration_id_to_slug(commonfields_id or integration_name)
        for cap_name in capabilities:
            cap_id = slugify_capability_name(cap_name)
            sub_cap_id = f"{cap_id}_{slug}"
            entries = cap_to_subcaps.setdefault(cap_id, [])
            if any(existing_id == sub_cap_id for existing_id, _ in entries):
                continue
            entries.append((sub_cap_id, license_values))

    records: list[tuple[str, str, list[str]]] = []
    for cap_id, entries in cap_to_subcaps.items():
        for sub_cap_id, license_values in entries:
            records.append((cap_id, sub_cap_id, license_values))
    return records


def write_csv(
    output_path: Path,
    license_columns: list[str],
    connector_to_records: dict[str, list[tuple[str, str, list[str]]]],
) -> None:
    """Write the grouped CSV with blank continuation cells."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(OUTPUT_LEADING_COLUMNS + license_columns)
        for connector_id, records in connector_to_records.items():
            first_connector_row = True
            last_cap_id: str | None = None
            for cap_id, sub_cap_id, license_values in records:
                connector_cell = connector_id if first_connector_row else ""
                cap_cell = cap_id if cap_id != last_cap_id else ""
                writer.writerow(
                    [connector_cell, cap_cell, sub_cap_id] + license_values
                )
                first_connector_row = False
                last_cap_id = cap_id


def write_xlsx(
    output_path: Path,
    license_columns: list[str],
    connector_to_records: dict[str, list[tuple[str, str, list[str]]]],
) -> None:
    """Write the grouped data to an XLSX with merged Connector ID / Capability
    cells.

    Connector ID cells are merged across all rows of a connector; Capability
    cells are merged across the consecutive rows that share a capability.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # a new Workbook always has an active sheet
    ws.title = "Licenses"

    header = OUTPUT_LEADING_COLUMNS + license_columns
    ws.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"
    )
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    center = Alignment(horizontal="center", vertical="center")

    current_row = 2  # row 1 is the header
    for connector_id, records in connector_to_records.items():
        connector_start = current_row
        # Track capability merge ranges within this connector.
        cap_start = current_row
        last_cap_id: str | None = None

        for cap_id, sub_cap_id, license_values in records:
            ws.cell(row=current_row, column=1, value=connector_id)
            ws.cell(row=current_row, column=2, value=cap_id)
            ws.cell(row=current_row, column=3, value=sub_cap_id)
            for offset, value in enumerate(license_values):
                ws.cell(row=current_row, column=4 + offset, value=value)

            if last_cap_id is None:
                last_cap_id = cap_id
                cap_start = current_row
            elif cap_id != last_cap_id:
                # Close the previous capability merge range.
                if current_row - 1 > cap_start:
                    ws.merge_cells(
                        start_row=cap_start,
                        start_column=2,
                        end_row=current_row - 1,
                        end_column=2,
                    )
                cap_start = current_row
                last_cap_id = cap_id

            current_row += 1

        # Close the final capability range of this connector.
        if current_row - 1 > cap_start:
            ws.merge_cells(
                start_row=cap_start,
                start_column=2,
                end_row=current_row - 1,
                end_column=2,
            )
        # Merge the connector cell across all its rows.
        if current_row - 1 > connector_start:
            ws.merge_cells(
                start_row=connector_start,
                start_column=1,
                end_row=current_row - 1,
                end_column=1,
            )

    # Center the merged columns for readability.
    for row_idx in range(2, current_row):
        ws.cell(row=row_idx, column=1).alignment = center
        ws.cell(row=row_idx, column=2).alignment = center

    # Reasonable column widths.
    widths = [22, 32, 40] + [16] * len(license_columns)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)


@main.command()
def aggregate(
    pipeline_csv: Path = typer.Option(
        Path("connectus/connectus-migration-pipeline.csv"),
        "--pipeline-csv",
        help="Path to the migration pipeline CSV.",
    ),
    license_csv: Path = typer.Option(
        Path(
            "connectus/integration_to_license_filtering.csv"
        ),
        "--license-csv",
        help="Path to the per-license filtering CSV.",
    ),
    output: Path = typer.Option(
        Path("connectus/connector_capability_licenses.csv"),
        "--output",
        "-o",
        help="Output CSV path. A sibling .xlsx is written alongside it.",
    ),
    repo_root: Path = typer.Option(
        Path("."),
        "--repo-root",
        help="Repo root used to resolve Integration File Path entries.",
    ),
) -> None:
    """Build the connector -> capability -> sub_capability license CSV + XLSX."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    license_columns, by_name = load_license_table(license_csv)
    pipeline_rows = read_pipeline_rows(pipeline_csv)
    grouped = group_by_connector(pipeline_rows)

    missing: list[str] = []
    unmatched_modules: set[str] = set()
    connector_to_records: dict[str, list[tuple[str, str, list[str]]]] = {}
    for connector_id, connector_rows in grouped.items():
        records = build_connector_records(
            connector_rows,
            by_name,
            license_columns,
            repo_root,
            missing,
            unmatched_modules,
        )
        if records:
            connector_to_records[connector_id] = records

    write_csv(output, license_columns, connector_to_records)
    xlsx_path = output.with_suffix(".xlsx")
    write_xlsx(xlsx_path, license_columns, connector_to_records)
    logger.info(f"[license_aggregator] Wrote {output} and {xlsx_path}")
    prase_licenses_to_json(Path("connectus/connector_capability_licenses.csv"))

    if unmatched_modules:
        logger.warning(
            f"[license_aggregator] {len(unmatched_modules)} supportedModules "
            f"value(s) did not match any license column:"
        )
        for module in sorted(unmatched_modules):
            logger.warning(f"  - {module}")

    if missing:
        unique_missing = sorted(set(missing))
        logger.warning(
            f"[license_aggregator] {len(unique_missing)} integration(s) had no "
            f"resolvable license and were excluded:"
        )
        for integration_id in unique_missing:
            logger.warning(f"  - {integration_id}")


if __name__ == "__main__":
    main()
